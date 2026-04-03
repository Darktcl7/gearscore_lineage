import openpyxl
import sys

wb = openpyxl.load_workbook(r'D:\Django Project\alto files\Power K9.xlsx', data_only=True)
for sn in wb.sheetnames:
    print(f'SHEET: {sn}')
    ws = wb[sn]
    print(f'Rows={ws.max_row} Cols={ws.max_column}')
    for r in range(1, min(40, ws.max_row+1)):
        row_data = []
        for c in range(1, min(ws.max_column+1, 30)):
            val = ws.cell(row=r, column=c).value
            if val is not None:
                row_data.append(f'{c}:{val}')
        print(f'R{r}| ' + ' | '.join(row_data))
    print('---END---')
sys.exit(0)

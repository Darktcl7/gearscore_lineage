import openpyxl

wb = openpyxl.load_workbook(r'D:\Django Project\alto files\Power K9.xlsx', data_only=False)
ws = wb.active
for i in range(2, 5):
    # D=4 (class), E=5 (level), F=6 (DMG), G=7 (ACC), H=8 (DEF), I=9 (DMG RED), J=10 (SKILL RES)
    # K=11 (SKILL DMG), L=12 (WPN DMG), M=13 (SS), N=14 (VALOR), O=15 (GUARD), P=16 (CONQ)
    # Q=17 (DUEL), R=18 (PURPLE), S=19 (GEAR SCORE)
    score_cell = ws.cell(row=i, column=19)
    print(f"Row {i} Name {ws.cell(row=i, column=2).value} Formula: {score_cell.value}")
